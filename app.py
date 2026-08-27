"""Imagine Inventory v2 — Imagine Events Tunisia
Gestion de depot + Emploi du temps + Export + Notifications"""
import os, uuid, json, io, hashlib
from datetime import datetime, date, timedelta, timezone
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy.exc import IntegrityError
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, Response, send_from_directory, has_request_context
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from PIL import Image, ImageOps, ImageFile
ImageFile.LOAD_TRUNCATED_IMAGES = True  # tolere les photos legerement tronquees
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Heure Tunis ──
TUNISIA_TZ = timezone(timedelta(hours=1))
def tunisia_now(): return datetime.now(TUNISIA_TZ)

# ── Config ──
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'imagine-events-tunisia-secret-2026')
# ── Cle du flux ICS (synchronisation Google/Apple/Outlook).
#    Derivee de SECRET_KEY : stable, unique par installation, pas de base de donnees. ──
ICS_KEY = hashlib.sha256((app.secret_key + ':ics-feed').encode('utf-8')).hexdigest()[:16]
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL:
    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///inventory.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# ── Anti-coupure : verifie que la connexion a la base est vivante avant chaque requete.
#    Render coupe parfois les connexions (surtout en pause d'inactivite). Sans ca, on a
#    des erreurs "SSL error: decryption failed or bad record mac". pool_pre_ping recree
#    automatiquement une connexion saine -> plus jamais de crash de ce type. ──
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,   # verifie la connexion avant usage
    'pool_recycle': 300,     # ferme les connexions de plus de 5 minutes
}
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Veuillez vous connecter.'

# ═══════════ M O D E L S ═══════════
class CustomRole(db.Model):
    __tablename__ = 'roles'
    id = db.Column(db.Integer, primary_key=True); name = db.Column(db.String(100), unique=True, nullable=False)
    icon = db.Column(db.String(10), default='👤'); description = db.Column(db.String(250), default='')
    permissions = db.Column(db.Text, default=''); created_at = db.Column(db.DateTime, default=tunisia_now)
    users = db.relationship('User', backref='role', lazy=True)
    def get_permissions(self):
        try: return json.loads(self.permissions) if self.permissions else []
        except: return []
    def set_permissions(self, pl): self.permissions = json.dumps(pl) if isinstance(pl, list) else '[]'
    def has_permission(self, p): return p in self.get_permissions()

class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True); email = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False); full_name = db.Column(db.String(150), nullable=False)
    role_id = db.Column(db.Integer, db.ForeignKey('roles.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=tunisia_now)
    borrows = db.relationship('Borrow', backref='user', lazy=True)
    event_assignments = db.relationship('EventAssignment', backref='user', lazy=True)
    notifications = db.relationship('Notification', backref='user', lazy=True, foreign_keys='Notification.user_id')
    def set_password(self, p): self.password_hash = generate_password_hash(p)
    def check_password(self, p): return check_password_hash(self.password_hash, p)
    def has_permission(self, perm):
        role = db.session.get(CustomRole, self.role_id) if self.role_id else None
        return role.has_permission(perm) if role else False
    @property
    def role_name(self):
        role = db.session.get(CustomRole, self.role_id) if self.role_id else None
        return role.name if role else 'Aucun role'
    @property
    def unread_notifications(self):
        return Notification.query.filter_by(user_id=self.id, read=False).count()

class Category(db.Model):
    __tablename__ = 'categories'
    id = db.Column(db.Integer, primary_key=True); name = db.Column(db.String(100), unique=True, nullable=False)
    icon = db.Column(db.String(10), default='📦')
    equipment = db.relationship('Equipment', backref='category', lazy=True)

class Equipment(db.Model):
    __tablename__ = 'equipment'
    id = db.Column(db.Integer, primary_key=True); name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, default=''); reference = db.Column(db.String(100), unique=True)
    category_id = db.Column(db.Integer, db.ForeignKey('categories.id'))
    total_quantity = db.Column(db.Integer, default=1); available_quantity = db.Column(db.Integer, default=1)
    in_repair = db.Column(db.Integer, default=0)  # unités en réparation
    specifications = db.Column(db.Text, default=''); condition = db.Column(db.String(50), default='Bon etat')
    location = db.Column(db.String(100), default='Depot principal'); created_at = db.Column(db.DateTime, default=tunisia_now)
    images = db.relationship('EquipmentImage', backref='equipment', lazy=True, cascade='all, delete-orphan')
    borrows = db.relationship('Borrow', backref='equipment', lazy=True)
    def primary_image(self): return self.images[0].filename if self.images else None
    def all_images(self): return [img.filename for img in self.images]

class EquipmentImage(db.Model):
    __tablename__ = 'equipment_images'
    id = db.Column(db.Integer, primary_key=True); filename = db.Column(db.String(300), nullable=False)
    equipment_id = db.Column(db.Integer, db.ForeignKey('equipment.id'), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=tunisia_now)

class ImageBlob(db.Model):
    """Copie PERMANENTE des photos dans la base de donnees.
    Sur Render, le dossier uploads/ est efface a CHAQUE mise a jour du site.
    On garde donc aussi les photos ici pour qu'elles ne disparaissent plus jamais."""
    __tablename__ = 'image_blobs'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(300), unique=True, nullable=False)
    data = db.Column(db.LargeBinary, nullable=False)
    mimetype = db.Column(db.String(50), default='image/jpeg')
    created_at = db.Column(db.DateTime, default=tunisia_now)

class Borrow(db.Model):
    __tablename__ = 'borrows'
    id = db.Column(db.Integer, primary_key=True); user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    equipment_id = db.Column(db.Integer, db.ForeignKey('equipment.id'), nullable=False)
    quantity = db.Column(db.Integer, default=1); borrow_date = db.Column(db.DateTime, default=tunisia_now)
    pickup_date = db.Column(db.Date, nullable=True)  # date de prise du materiel
    expected_return_date = db.Column(db.Date, nullable=False); actual_return_date = db.Column(db.DateTime, nullable=True)
    status = db.Column(db.String(50), default='active'); notes = db.Column(db.Text, default='')
    event_name = db.Column(db.String(200), default='')
    event_id = db.Column(db.Integer, db.ForeignKey('events.id'), nullable=True)  # lien vers un evenement

class ActivityLog(db.Model):
    __tablename__ = 'activity_logs'
    id = db.Column(db.Integer, primary_key=True); user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    action = db.Column(db.String(50), nullable=False); description = db.Column(db.Text, default='')
    equipment_name = db.Column(db.String(200), default=''); quantity = db.Column(db.Integer, default=0)
    timestamp = db.Column(db.DateTime, default=tunisia_now)

# ── NEW: Events (emploi du temps) ──
class Event(db.Model):
    __tablename__ = 'events'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, default='')
    event_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=True)  # evenement multi-jours : derniere journee (NULL = 1 jour)
    start_time = db.Column(db.String(10), default='08:00')  # HH:MM
    end_time = db.Column(db.String(10), default='17:00')
    location = db.Column(db.String(200), default='')
    status = db.Column(db.String(50), default='upcoming')  # upcoming / ongoing / completed / cancelled
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=tunisia_now)
    assignments = db.relationship('EventAssignment', backref='event', lazy=True, cascade='all, delete-orphan')
    def date_end(self):
        return self.end_date or self.event_date
    def is_multiday(self):
        return bool(self.end_date) and self.end_date != self.event_date
    def date_label(self):
        if self.is_multiday():
            return f"du {self.event_date.strftime('%d/%m/%Y')} au {self.date_end().strftime('%d/%m/%Y')}"
        return f"le {self.event_date.strftime('%d/%m/%Y')}"

class EventAssignment(db.Model):
    __tablename__ = 'event_assignments'
    id = db.Column(db.Integer, primary_key=True)
    event_id = db.Column(db.Integer, db.ForeignKey('events.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    role = db.Column(db.String(100), default='Staff')  # Staff, Responsable, Technicien, etc.
    notes = db.Column(db.Text, default='')

# ── NEW: Notifications ──
class Notification(db.Model):
    __tablename__ = 'notifications'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text, default='')
    link = db.Column(db.String(300), default='')
    read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=tunisia_now)

# ── NEW: Commandes de materiel personnalise (notes) ──
class MaterialOrder(db.Model):
    __tablename__ = 'material_orders'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, default='')
    status = db.Column(db.String(50), default='nouvelle')   # nouvelle / commandee / recue
    priority = db.Column(db.String(20), default='normale')   # basse / normale / haute
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=tunisia_now)
    updated_at = db.Column(db.DateTime, default=tunisia_now, onupdate=tunisia_now)

# ═══════════ H E L P E R S ═══════════
@login_manager.user_loader
def load_user(uid): return db.session.get(User, int(uid))

@app.template_filter('get_user')
def get_user_filter(uid): return db.session.get(User, int(uid)) if uid else None

def allowed_file(fn): return '.' in fn and fn.rsplit('.',1)[1].lower() in ALLOWED_EXTENSIONS

def save_uploaded_image(file):
    """Traite une photo le plus vite possible :
    - lit directement le fichier en memoire (pas d'ecriture disque inutile)
    - mode 'draft' pour les JPEG = decodage beaucoup plus rapide des grosses photos
    - redimensionne a 1000px max (suffisant pour l'affichage, mais 4 a 10x plus leger)
    - corrige l'orientation des photos iPhone (EXIF)
    Puis garde une copie PERMANENTE dans la base (image_blobs)."""
    if not file or not file.filename: return None
    if not allowed_file(file.filename): return None
    file.seek(0, 2); size = file.tell(); file.seek(0)
    if size == 0: return None
    ext = file.filename.rsplit('.', 1)[1].lower()
    jname = f"{uuid.uuid4().hex}.jpg"
    jpath = os.path.join(app.config['UPLOAD_FOLDER'], jname)
    try:
        data = file.read()
        img = Image.open(io.BytesIO(data))
        # draft : decode JPEG a resolution reduite -> 3 a 5x plus rapide
        if img.format == 'JPEG':
            try: img.draft('RGB', (2000, 2000))
            except Exception: pass
        img = ImageOps.exif_transpose(img)          # orientation iPhone
        img.thumbnail((1000, 1000))                  # taille finale
        if img.mode in ('RGBA', 'P', 'LA'): img = img.convert('RGB')
        buf = io.BytesIO()
        img.save(buf, 'JPEG', optimize=True, quality=82)
        blob = buf.getvalue()
        with open(jpath, 'wb') as fh:
            fh.write(blob)
        # Copie PERMANENTE dans la base de donnees
        try:
            db.session.add(ImageBlob(filename=jname, data=blob, mimetype='image/jpeg'))
        except Exception as e:
            print(f'[IMG-DB] impossible de stocker {jname} en base: {e}')
        return jname
    except Exception as e:
        print(f'[IMG] conversion impossible ({e}), fichier original conserve')
        # Fallback : garder le fichier original tel quel
        try:
            file.seek(0)
            uname = f"{uuid.uuid4().hex}.{ext}"
            fpath = os.path.join(app.config['UPLOAD_FOLDER'], uname)
            file.save(fpath)
            return uname
        except Exception:
            return None

def borrow_out_on(b, d):
    """Un emprunt est 'sorti' au jour d si la periode [prise, retour prevu] couvre d.
    - avant la prise : pas encore sorti
    - du jour de la prise jusqu'a AUJOURD'HUI : sorti (le materiel est en main,
      meme si l'emprunt est en retard : retour non fait)
    - jours a venir : sorti jusqu'a la date de retour prevue"""
    start = b.pickup_date or (b.borrow_date.date() if b.borrow_date else None)
    end = b.expected_return_date
    if not start or not end:
        return False
    if d < start:
        return False
    if d <= date.today():
        return True
    return d <= end


def units_out_on(eq_id, d):
    """Nombre d'unites emportees/sorties le jour d (emprunts actifs/en retard)."""
    total = 0
    for b in Borrow.query.filter_by(equipment_id=eq_id).filter(Borrow.status.in_(['active', 'late'])).all():
        if borrow_out_on(b, d):
            total += b.quantity or 0
    return total


def availability_on_date(eq, d):
    """Unites disponibles du jour d = total - reparations - emprunts sortis ce jour-la."""
    return max(0, (eq.total_quantity or 0) - (eq.in_repair or 0) - units_out_on(eq.id, d))


def min_availability_range(eq, start, end, max_days=120):
    """La disponibilite MINIMUM sur toute la periode [start, end] (jour par jour).
    Sert a empecher de double-emprunter une periode deja reservee."""
    borrows = Borrow.query.filter_by(equipment_id=eq.id).filter(Borrow.status.in_(['active', 'late'])).all()
    d = start
    mn = None
    for _ in range(max_days):
        if d > end:
            break
        out = sum((b.quantity or 0) for b in borrows if borrow_out_on(b, d))
        a = max(0, (eq.total_quantity or 0) - (eq.in_repair or 0) - out)
        mn = a if mn is None else min(mn, a)
        d += timedelta(days=1)
    if mn is None:
        mn = max(0, (eq.total_quantity or 0) - (eq.in_repair or 0))
    return mn


def update_availability(eq_id):
    eq = db.session.get(Equipment, eq_id)
    if eq:
        eq.available_quantity = availability_on_date(eq, date.today()); db.session.commit()


def is_pending_user():
    """Verifie si l'utilisateur est en attente (aucune permission)"""
    return not current_user.has_permission('borrow_equipment') and not current_user.has_permission('manage_equipment') and not current_user.has_permission('manage_users')

def log_action(action, description='', equipment_name='', quantity=0):
    """Enregistre une action. Si l'utilisateur n'est pas connecte, user_id = None."""
    uid = current_user.id if current_user.is_authenticated else None
    log = ActivityLog(user_id=uid, action=action, description=description, equipment_name=equipment_name, quantity=quantity)
    db.session.add(log); db.session.commit()

def notify_user(uid, title, message, link=''):
    n = Notification(user_id=uid, title=title, message=message, link=link)
    db.session.add(n); db.session.commit()
    # Envoi email si cle API configuree
    try:
        u = db.session.get(User, uid)
        if u and os.environ.get('SMTP_PASSWORD'):
            html = email_template(title, f'Bonjour {u.full_name},', message, absolute_link(link), 'Voir dans Imagine Inventory')
            send_email(u.email, title, html, message)
    except Exception as e:
        print(f'[EMAIL ERROR] {type(e).__name__}: {e}')


def send_email(to, subject, body_html, body_text=''):
    """Envoi email via Brevo API HTTP avec template HTML"""
    import urllib.request, json as jmod
    api_key = os.environ.get('SMTP_PASSWORD','')
    if not api_key: return
    sender_email = os.environ.get('SMTP_FROM','')
    if not sender_email: return
    data = jmod.dumps({
        "sender": {"name": "Imagine Events Tunisia", "email": sender_email},
        "to": [{"email": to}],
        "subject": f'[Imagine Inventory] {subject}',
        "htmlContent": body_html,
        "textContent": body_text or subject
    }).encode('utf-8')
    req = urllib.request.Request(
        'https://api.brevo.com/v3/smtp/email',
        data=data,
        headers={'accept':'application/json','api-key':api_key,'content-type':'application/json'},
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            print(f'[EMAIL] OK -> {to}')
    except urllib.error.HTTPError as e:
        err_body = e.read().decode()[:200] if e.fp else 'N/A'
        print(f'[EMAIL] HTTP {e.code}: {err_body}')
    except Exception as e:
        print(f'[EMAIL] Error: {e}')


def absolute_link(path):
    """Transforme un lien relatif (/dashboard) en lien COMPLET (https://site.com/dashboard).
    Les emails doivent contenir des liens complets : un lien relatif dans un email
    donne une erreur 404 quand on clique dessus."""
    if not path:
        return ''
    if path.startswith('http'):
        return path
    site = os.environ.get('SITE_URL', '').strip().rstrip('/')
    if not site:
        try:
            if has_request_context():
                site = request.host_url.rstrip('/')
                # En production (Render/OVH), forcer https (le site est toujours en https)
                if not site.startswith('http://localhost') and not site.startswith('http://127.0.0.1'):
                    site = site.replace('http://', 'https://', 1)
        except Exception:
            site = ''
    if site:
        return site + '/' + path.lstrip('/')
    return path


def email_template(title, greeting, content, action_link='', action_text=''):
    """Template HTML pour les emails Imagine Events"""
    action_btn = f'<a href="{action_link}" style="display:inline-block;background:#C41E3A;color:white;padding:12px 30px;border-radius:6px;text-decoration:none;font-weight:bold;margin:15px 0;font-size:14px">{action_text}</a>' if action_link else ''
    return f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;font-family:Arial,Helvetica,sans-serif;background:#F8FAFC">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#F8FAFC;padding:20px">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="background:white;border-radius:12px;overflow:hidden;box-shadow:0 2px 10px rgba(0,0,0,.08)">
<tr><td style="background:linear-gradient(135deg,#0B1D3A,#13294B);padding:25px 30px;text-align:center">
    <div style="font-size:32px;margin-bottom:8px">\u2726</div>
    <div style="font-family:Georgia,serif;font-size:20px;font-weight:bold;color:white;letter-spacing:2px">IMAGINE<span style="color:#E63946"> EVENTS</span></div>
    <div style="font-size:11px;color:rgba(255,255,255,.5);letter-spacing:2px;margin-top:4px">TUNISIA · EXCELLENCE EVENEMENTIELLE</div>
</td></tr>
<tr><td style="padding:30px">
    <h2 style="color:#0B1D3A;font-size:18px;margin:0 0 10px">{title}</h2>
    <p style="color:#64748B;font-size:14px;line-height:1.6;margin:0 0 15px">{greeting}</p>
    <div style="background:#F1F5F9;border-left:4px solid #C41E3A;padding:15px;border-radius:0 8px 8px 0;margin:15px 0">
        <p style="color:#334155;font-size:14px;margin:0;line-height:1.6">{content}</p>
    </div>
    {action_btn}
    <p style="color:#94A3B8;font-size:12px;margin-top:20px">Cet email a ete envoye automatiquement par Imagine Inventory.</p>
</td></tr>
<tr><td style="background:#0B1D3A;padding:15px 30px;text-align:center">
    <p style="color:rgba(255,255,255,.5);font-size:11px;margin:0">
        Rue du Lac Loch Ness, Imm Neo, 3eme etage, 1053 Les Berges du Lac, Tunis<br>
        +216 71 656 056 | info@imagine-events.com
    </p>
</td></tr>
</table>
</td></tr></table>
</body></html>'''


def notify_admins(title, message, link=''):
    admin_role = CustomRole.query.filter_by(name='Admin').first()
    if admin_role:
        for a in User.query.filter_by(role_id=admin_role.id).all():
            notify_user(a.id, title, message, link)

ALL_PERMISSIONS = [
    {"key":"manage_users","label":"Gerer les utilisateurs","desc":"Creer, modifier, supprimer des comptes","icon":"👥"},
    {"key":"manage_roles","label":"Gerer les roles","desc":"Creer et modifier les roles et permissions","icon":"🔐"},
    {"key":"manage_equipment","label":"Gerer le materiel","desc":"Ajouter, modifier, supprimer du materiel","icon":"📦"},
    {"key":"borrow_equipment","label":"Emprunter","desc":"Emprunter du materiel","icon":"📤"},
    {"key":"return_equipment","label":"Retourner","desc":"Marquer un emprunt comme retourne","icon":"✅"},
    {"key":"manage_categories","label":"Gerer les categories","desc":"Ajouter des categories de materiel","icon":"📂"},
    {"key":"view_logs","label":"Voir les logs","desc":"Consulter l'historique d'activite","icon":"📜"},
    {"key":"clear_history","label":"Effacer l'historique","desc":"Supprimer l'historique des emprunts","icon":"🗑️"},
    {"key":"manage_schedule","label":"Gerer le planning","desc":"Creer et gerer l'emploi du temps","icon":"📅"},
    {"key":"manage_database","label":"Gerer la base de donnees","desc":"Reset, import CSV, restauration photos, migration","icon":"🗄️"},
]

def permission_required(perm):
    def decorator(f):
        @wraps(f)
        @login_required
        def decorated(*args, **kwargs):
            if not current_user.has_permission(perm):
                try: log_action('permission_denied', f'Tentative d\'acces refuse : {perm}')
                except: pass
                flash('Acces refuse. Permission requise : ' + perm, 'error')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated
    return decorator

# ── Page d'erreur propre (au lieu d'un crash brut) ──
@app.errorhandler(500)
def handle_internal_error(e):
    try:
        db.session.rollback()
    except Exception:
        pass
    try:
        db.session.remove()
    except Exception:
        pass
    print(f'[ERREUR 500] {type(e).__name__}: {e}')
    return render_template('error.html'), 500

# ═══════════ R O U T E S ═══════════
@app.route('/')
def index():
    return redirect(url_for('dashboard') if current_user.is_authenticated else url_for('login'))

@app.route('/login', methods=['GET','POST'])
def login():
    if current_user.is_authenticated: return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email','').strip().lower(); pw = request.form.get('password','')
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(pw):
            login_user(user, remember=True)
            db.session.add(ActivityLog(user_id=user.id, action='login', description=f'Connexion de {user.full_name}')); db.session.commit()
            flash(f'Bienvenue {user.full_name} !', 'success')
            return redirect(request.args.get('next') or url_for('dashboard'))
        log_action('failed_login', f'Tentative echouee pour {email}'); flash('Email ou mot de passe incorrect.', 'error')
    return render_template('login.html')

@app.route('/register', methods=['GET','POST'])
def register():
    if current_user.is_authenticated: return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email','').strip().lower(); name = request.form.get('full_name','').strip()
        pw = request.form.get('password','')
        if not email or not name or not pw: flash('Tous les champs requis.','error'); return render_template('register.html')
        elif request.form.get('confirm_password') != pw: flash('Mots de passe differents.','error'); return render_template('register.html')
        elif len(pw) < 6: flash('6 caracteres minimum.','error'); return render_template('register.html')
        elif User.query.filter_by(email=email).first(): flash('Email deja utilise.','error'); return render_template('register.html')
        else:
            try:
                u = User(email=email, full_name=name); u.set_password(pw)
                if User.query.count() == 0: u.role_id = 1
                else:
                    pending = CustomRole.query.filter_by(name='En attente').first()
                    u.role_id = pending.id if pending else None
                db.session.add(u); db.session.commit()
                # Notify admins
                notify_admins(f'Nouveau compte : {name}', f'{name} ({email}) vient de creer un compte. Action requise.', '/admin/users')
                log_action('register', f'Nouveau compte cree par {name} ({email})'); flash('Compte cree ! En attente de validation par l\'admin.','success')
                return redirect(url_for('login'))
            except Exception as e:
                db.session.rollback()
                flash('Erreur lors de la creation du compte. Reessayez.','error')
                print(f'[REGISTER ERROR] {e}')
                return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    log_action('logout', f'Deconnexion de {current_user.full_name}')
    logout_user()
    flash('Vous etes deconnecte.','info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    search = request.args.get('search','').strip(); cat_filter = request.args.get('category','')
    q = Equipment.query
    if search: q = q.filter(db.or_(Equipment.name.ilike(f'%{search}%'),Equipment.description.ilike(f'%{search}%'),Equipment.reference.ilike(f'%{search}%'),Equipment.specifications.ilike(f'%{search}%'),Equipment.location.ilike(f'%{search}%')))
    if cat_filter: q = q.filter_by(category_id=int(cat_filter))
    eqs = q.order_by(Equipment.name).all(); cats = Category.query.order_by(Category.name).all()
    ab = Borrow.query.filter(Borrow.status.in_(['active','late'])).order_by(Borrow.expected_return_date).all()
    today = date.today()
    for b in ab:
        if b.status == 'active' and b.expected_return_date < today: b.status = 'late'
    db.session.commit()
    eq_json = json.dumps([{'id':e.id,'name':e.name,'available_quantity':e.available_quantity} for e in Equipment.query.all()])
    is_pending = not current_user.has_permission('borrow_equipment') and not current_user.has_permission('manage_equipment') and not current_user.has_permission('manage_users')
    upcoming_events = Event.query.filter(Event.end_date >= date.today()).order_by(Event.event_date, Event.start_time).all()
    repair_count = Equipment.query.filter(Equipment.in_repair > 0).count()
    return render_template('dashboard.html', equipment=eqs, categories=cats, active_borrows=ab, search=search, cat_filter=cat_filter, all_count=Equipment.query.count(), available_count=Equipment.query.filter(Equipment.available_quantity>0).count(), late_count=Borrow.query.filter_by(status='late').count(), repair_count=repair_count, equipment_json=eq_json, today=today, is_pending=is_pending, upcoming_events=upcoming_events)

@app.route('/equipment/<int:eid>')
@login_required
def equipment_detail(eid):
    if is_pending_user(): flash('Votre compte est en attente de validation.','error'); return redirect(url_for('dashboard'))
    eq = db.session.get(Equipment, eid)
    if not eq: flash('Introuvable.','error'); return redirect(url_for('dashboard'))
    borrows = Borrow.query.filter_by(equipment_id=eid).order_by(Borrow.borrow_date.desc()).all()
    upcoming_events = Event.query.filter(Event.end_date >= date.today()).order_by(Event.event_date, Event.start_time).all()
    # ── Donnees du calendrier de disponibilite (emprunts en cours / en retard) ──
    cal_borrows = []
    for b in Borrow.query.filter_by(equipment_id=eid).filter(Borrow.status.in_(['active', 'late'])).order_by(Borrow.pickup_date, Borrow.id).all():
        start = b.pickup_date or (b.borrow_date.date() if b.borrow_date else None)
        end = b.expected_return_date
        if not (start and end):
            continue
        cal_borrows.append({
            'id': b.id,
            'user': b.user.full_name if b.user else '?',
            'qty': b.quantity or 0,
            'start': start.strftime('%Y-%m-%d'),
            'end': end.strftime('%Y-%m-%d'),
            'start_fmt': start.strftime('%d/%m/%Y'),
            'end_fmt': end.strftime('%d/%m/%Y'),
            'status': b.status,
            'event': b.event_name or ''
        })
    return render_template('equipment_detail.html', eq=eq, borrows=borrows, today=date.today(), upcoming_events=upcoming_events, cal_borrows=cal_borrows, cal_borrows_json=json.dumps(cal_borrows), today_str=date.today().strftime('%Y-%m-%d'))

@app.route('/borrow/<int:eid>', methods=['POST'])
@permission_required('borrow_equipment')
def borrow_equipment(eid):
    eq = db.session.get(Equipment, eid)
    if not eq: flash('Introuvable.','error'); return redirect(url_for('dashboard'))
    qty = int(request.form.get('quantity',1)); rd = request.form.get('return_date','')
    try: return_date = datetime.strptime(rd,'%Y-%m-%d').date()
    except: flash('Date invalide.','error'); return redirect(url_for('dashboard'))
    if return_date < date.today(): flash('Date dans le passe.','error'); return redirect(url_for('dashboard'))
    # ── Date de prise du materiel (optionnelle, defaut = aujourd'hui) ──
    pd = request.form.get('pickup_date','')
    try:
        pickup_date = datetime.strptime(pd,'%Y-%m-%d').date() if pd else date.today()
    except (TypeError, ValueError):
        pickup_date = date.today()
    if return_date < pickup_date:
        flash('La date de retour doit etre apres la date de prise.','error'); return redirect(url_for('dashboard'))
    # ── Quantite : verifiee sur TOUTE la periode [prise, retour] ──
    # Un emprunt reserve le materiel sur sa periode : impossible de double-
    # emprunter (ex: materiel deja reserve pour un evenement confirme).
    if qty < 1:
        flash('Quantite invalide.','error'); return redirect(url_for('dashboard'))
    min_avail = min_availability_range(eq, pickup_date, return_date)
    if qty > min_avail:
        flash(f'Quantite invalide : il ne reste que {min_avail} unite(s) disponible(s) du {pickup_date.strftime("%d/%m")} au {return_date.strftime("%d/%m")} (emprunts en cours compris).','error')
        return redirect(url_for('dashboard'))
    # ── Evenement : soit choisi dans la liste (event_id), soit ecrit librement ──
    evt = None
    try:
        eid_form = int(request.form.get('event_id','') or 0)
        if eid_form: evt = db.session.get(Event, eid_form)
    except (TypeError, ValueError):
        evt = None
    event_name = request.form.get('event_name','').strip()
    if evt: event_name = evt.title  # priorite a l'evenement choisi
    b = Borrow(user_id=current_user.id, equipment_id=eid, quantity=qty, expected_return_date=return_date, pickup_date=pickup_date, event_name=event_name, event_id=evt.id if evt else None, notes=request.form.get('notes','').strip())
    db.session.add(b); db.session.commit(); update_availability(eid)
    log_action('borrow', f'Emprunt de {qty}x {eq.name}', eq.name, qty)
    flash(f'{qty} x {eq.name} emprunte(s). Prise le {pickup_date.strftime("%d/%m/%Y")}, retour le {return_date.strftime("%d/%m/%Y")}.','success')
    return redirect(url_for('dashboard'))

@app.route('/return/<int:bid>', methods=['POST'])
@permission_required('return_equipment')
def return_equipment(bid):
    b = db.session.get(Borrow, bid)
    if not b or b.status not in ('active','late'): flash('Emprunt introuvable ou deja retourne.','error'); return redirect(url_for('dashboard'))
    is_admin = current_user.has_permission('manage_users')
    if b.user_id != current_user.id and not is_admin: flash('Vous ne pouvez retourner que vos propres emprunts.','error'); return redirect(url_for('dashboard'))
    # ── Combien partent en reparation, combien sont propres ? ──
    try:
        repair_qty = int(request.form.get('repair_qty','') or 0)
    except (TypeError, ValueError):
        repair_qty = 0
    total = b.quantity
    if repair_qty < 0 or repair_qty > total: repair_qty = 0
    ready_qty = total - repair_qty
    eq = db.session.get(Equipment, b.equipment_id)
    if eq and repair_qty > 0:
        eq.in_repair = (eq.in_repair or 0) + repair_qty
    b.status = 'returned'; b.actual_return_date = tunisia_now(); db.session.commit(); update_availability(b.equipment_id)
    who = current_user.full_name if b.user_id == current_user.id else f'{current_user.full_name} (admin) pour {b.user.full_name}'
    log_action('return', f'Retour de {b.quantity}x {b.equipment.name} par {who} ({ready_qty} propres, {repair_qty} en reparation)', b.equipment.name, b.quantity)
    flash(f'{b.equipment.name} retourne : {ready_qty} propre(s) et pret(s), {repair_qty} en reparation.','success')
    return redirect(url_for('dashboard'))

@app.route('/equipment/<int:eid>/repair', methods=['POST'])
@permission_required('manage_equipment')
def mark_repaired(eid):
    """Remettre en stock des unites qui etaient en reparation."""
    eq = db.session.get(Equipment, eid)
    if not eq: flash('Introuvable.','error'); return redirect(url_for('dashboard'))
    try:
        qty = int(request.form.get('qty','') or 0)
    except (TypeError, ValueError):
        qty = 0
    if qty <= 0 or qty > (eq.in_repair or 0):
        flash(f'Quantite invalide (max {eq.in_repair or 0} en reparation).','error')
        return redirect(url_for('equipment_detail', eid=eid))
    eq.in_repair -= qty
    db.session.commit(); update_availability(eid)
    log_action('return', f'{qty}x {eq.name} repare(s) et remis en stock', eq.name, qty)
    flash(f'{qty} unite(s) de {eq.name} reparee(s) et remise(s) en stock.','success')
    return redirect(url_for('equipment_detail', eid=eid))

# ═══════ E M P R U N T S  (preparation par evenement) ═══════
@app.route('/borrows')
@login_required
def borrows_page():
    if is_pending_user(): flash('Votre compte est en attente de validation.','error'); return redirect(url_for('dashboard'))
    active = Borrow.query.filter(Borrow.status.in_(['active','late'])).order_by(Borrow.expected_return_date).all()
    events = Event.query.filter(Event.end_date >= date.today()).order_by(Event.event_date, Event.start_time).all()
    # Grouper par evenement
    by_event = {}   # event_id -> {'event': evt, 'borrows': [...]}
    no_event = []
    for b in active:
        if b.event_id:
            grp = by_event.setdefault(b.event_id, {'event': db.session.get(Event, b.event_id), 'borrows': []})
            grp['borrows'].append(b)
        else:
            no_event.append(b)
    repair_list = Equipment.query.filter(Equipment.in_repair > 0).order_by(Equipment.name).all()
    return render_template('borrows.html', by_event=by_event, no_event=no_event, events=events, repair_list=repair_list, today=date.today(), active_count=len(active))

# ═══════ R E P A R A T I O N S ═══════
@app.route('/repairs')
@login_required
def repairs_page():
    if is_pending_user(): flash('Votre compte est en attente de validation.','error'); return redirect(url_for('dashboard'))
    repair_list = Equipment.query.filter(Equipment.in_repair > 0).order_by(Equipment.name).all()
    total_repair = sum((e.in_repair or 0) for e in repair_list)
    return render_template('repairs.html', repair_list=repair_list, total_repair=total_repair, today=date.today())

@app.route('/borrows/<int:bid>/assign-event', methods=['POST'])
@login_required
def assign_event_to_borrow(bid):
    if is_pending_user(): flash('Votre compte est en attente de validation.','error'); return redirect(url_for('dashboard'))
    if not current_user.has_permission('borrow_equipment') and not current_user.has_permission('manage_schedule'):
        flash('Acces refuse.','error'); return redirect(url_for('borrows_page'))
    b = db.session.get(Borrow, bid)
    if not b: flash('Emprunt introuvable.','error'); return redirect(url_for('borrows_page'))
    event_id = request.form.get('event_id','')
    if event_id:
        evt = db.session.get(Event, int(event_id))
        if evt:
            b.event_id = evt.id; b.event_name = evt.title
            flash(f'Emprunt associe a l\'evenement "{evt.title}".','success')
    else:
        b.event_id = None; b.event_name = ''
        flash('Emprunt dissocie de tout evenement.','info')
    db.session.commit()
    return redirect(url_for('borrows_page'))

# ═══════ C O M M A N D E S  (materiel personnalise) ═══════
@app.route('/orders')
@login_required
def orders_page():
    if is_pending_user(): flash('Votre compte est en attente de validation.','error'); return redirect(url_for('dashboard'))
    orders = MaterialOrder.query.order_by(MaterialOrder.created_at.desc()).all()
    return render_template('orders.html', orders=orders)

@app.route('/orders/create', methods=['POST'])
@permission_required('manage_equipment')
def create_order():
    title = request.form.get('title','').strip()
    if not title: flash('Le titre est requis.','error'); return redirect(url_for('orders_page'))
    o = MaterialOrder(title=title, description=request.form.get('description','').strip(), priority=request.form.get('priority','normale'), created_by=current_user.id)
    db.session.add(o); db.session.commit()
    log_action('create_order', f'Commande creee : {title}')
    flash(f'Commande "{title}" ajoutee.','success')
    return redirect(url_for('orders_page'))

@app.route('/orders/<int:oid>/status', methods=['POST'])
@permission_required('manage_equipment')
def set_order_status(oid):
    o = db.session.get(MaterialOrder, oid)
    if not o: flash('Commande introuvable.','error'); return redirect(url_for('orders_page'))
    st = request.form.get('status','')
    if st in ('nouvelle','commandee','recue'):
        o.status = st
        db.session.commit()
        labels = {'nouvelle':'nouvelle','commandee':'commandee','recue':'recue'}
        flash(f'Commande "{o.title}" -> {labels[st]}.','success')
    return redirect(url_for('orders_page'))

@app.route('/orders/<int:oid>/delete', methods=['POST'])
@permission_required('manage_equipment')
def delete_order(oid):
    o = db.session.get(MaterialOrder, oid)
    if o:
        db.session.delete(o); db.session.commit()
        flash(f'Commande "{o.title}" supprimee.','info')
    return redirect(url_for('orders_page'))

@app.route('/equipment/add', methods=['GET','POST'])
@permission_required('manage_equipment')
def add_equipment():
    if request.method == 'POST':
        name = request.form.get('name','').strip()
        if not name: flash('Nom requis.','error'); return redirect(url_for('add_equipment'))
        ref = request.form.get('reference','').strip() or f"IM-{uuid.uuid4().hex[:8].upper()}"
        # ── Verification AVANT d'enregistrer : la reference doit etre unique ──
        if Equipment.query.filter(db.func.lower(Equipment.reference) == ref.lower()).first():
            flash(f'⚠️ La reference "{ref}" existe deja. Choisis une autre reference (ou laisse le champ vide pour en generer une automatiquement).','error')
            return redirect(url_for('add_equipment'))
        # ── Conversion en toute securite (ne plante jamais si champ vide) ──
        try:
            cid = int(request.form.get('category_id','') or 0) or None
        except (TypeError, ValueError):
            cid = None
        try:
            total_qty = int(request.form.get('total_quantity','') or 1)
        except (TypeError, ValueError):
            total_qty = 1
        if total_qty < 1: total_qty = 1
        try:
            eq = Equipment(name=name, description=request.form.get('description','').strip(), reference=ref, category_id=cid, total_quantity=total_qty, available_quantity=total_qty, specifications=request.form.get('specifications','').strip(), condition=request.form.get('condition','Bon etat'), location=request.form.get('location','Depot principal'))
            db.session.add(eq); db.session.flush()
            for f in request.files.getlist('images'):
                sn = save_uploaded_image(f)
                if sn: db.session.add(EquipmentImage(filename=sn, equipment_id=eq.id))
            db.session.commit()
            log_action('add_equipment', f'Ajout de {name}', name)
            flash(f'"{name}" ajoute au depot !','success')
            return redirect(url_for('equipment_detail', eid=eq.id))
        except IntegrityError:
            db.session.rollback()
            flash('⚠️ Cette reference est deja utilisee par un autre materiel. Choisis une autre reference.','error')
            return redirect(url_for('add_equipment'))
        except Exception as e:
            db.session.rollback()
            flash('Erreur lors de l\'ajout. Reessaie.','error')
            print(f'[ADD EQUIPMENT ERROR] {type(e).__name__}: {e}')
            return redirect(url_for('add_equipment'))
    return render_template('add_equipment.html', categories=Category.query.order_by(Category.name).all())

@app.route('/equipment/<int:eid>/edit', methods=['GET','POST'])
@permission_required('manage_equipment')
def edit_equipment(eid):
    eq = db.session.get(Equipment, eid)
    if not eq: flash('Introuvable.','error'); return redirect(url_for('dashboard'))
    if request.method == 'POST':
        eq.name = request.form.get('name','').strip(); eq.description = request.form.get('description','').strip()
        ref = request.form.get('reference','').strip()
        if ref and ref.lower() != (eq.reference or '').lower():
            # ── La reference doit rester unique : verifier qu'aucun AUTRE materiel ne l'a ──
            autre = Equipment.query.filter(db.func.lower(Equipment.reference) == ref.lower(), Equipment.id != eid).first()
            if autre:
                flash(f'⚠️ La reference "{ref}" est deja utilisee par "{autre.name}". Choisis une autre reference.','error')
                return redirect(url_for('edit_equipment', eid=eid))
            eq.reference = ref
        cid = request.form.get('category_id')
        try: eq.category_id = int(cid) if cid else None
        except (TypeError, ValueError): eq.category_id = None
        old = eq.total_quantity
        try: eq.total_quantity = int(request.form.get('total_quantity', eq.total_quantity) or eq.total_quantity)
        except (TypeError, ValueError): pass
        eq.available_quantity = max(0, eq.available_quantity + (eq.total_quantity - old))
        eq.specifications = request.form.get('specifications','').strip(); eq.condition = request.form.get('condition','Bon etat'); eq.location = request.form.get('location','Depot principal')
        for f in request.files.getlist('images'):
            sn = save_uploaded_image(f)
            if sn: db.session.add(EquipmentImage(filename=sn, equipment_id=eq.id))
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash('⚠️ Cette reference est deja utilisee par un autre materiel. Choisis une autre reference.','error')
            return redirect(url_for('edit_equipment', eid=eid))
        log_action('edit_equipment', f'Modification de {eq.name}', eq.name)
        flash(f'"{eq.name}" mis a jour.','success')
        return redirect(url_for('equipment_detail', eid=eq.id))
    return render_template('edit_equipment.html', eq=eq, categories=Category.query.order_by(Category.name).all())

@app.route('/equipment/<int:eid>/delete-image/<int:iid>', methods=['POST'])
@login_required
def delete_image(eid, iid):
    img = db.session.get(EquipmentImage, iid)
    if img and img.equipment_id == eid:
        fp = os.path.join(app.config['UPLOAD_FOLDER'], img.filename)
        if os.path.exists(fp): os.remove(fp)
        blob = ImageBlob.query.filter_by(filename=img.filename).first()
        if blob: db.session.delete(blob)
        db.session.delete(img); db.session.commit()
    return redirect(url_for('edit_equipment', eid=eid))

@app.route('/equipment/<int:eid>/delete', methods=['POST'])
@permission_required('manage_equipment')
def delete_equipment(eid):
    eq = db.session.get(Equipment, eid)
    if eq:
        for img in eq.images:
            fp = os.path.join(app.config['UPLOAD_FOLDER'], img.filename)
            if os.path.exists(fp): os.remove(fp)
            blob = ImageBlob.query.filter_by(filename=img.filename).first()
            if blob: db.session.delete(blob)
        db.session.delete(eq); db.session.commit()
        log_action('delete_equipment', f'Suppression de {eq.name}', eq.name)
        flash(f'"{eq.name}" supprime.','info')
    return redirect(url_for('dashboard'))

@app.route('/uploads/<fn>')
@login_required
def uploaded_file(fn):
    # 1) D'abord la copie PERMANENTE en base de donnees
    blob = ImageBlob.query.filter_by(filename=fn).first()
    if blob and blob.data:
        return Response(blob.data, mimetype=blob.mimetype or 'image/jpeg')
    # 2) Sinon, le fichier sur le disque (anciennes photos non encore migrees)
    return send_from_directory(app.config['UPLOAD_FOLDER'], fn)

@app.route('/change-password', methods=['GET','POST'])
@login_required
def change_password():
    if request.method == 'POST':
        cp = request.form.get('current_password',''); np = request.form.get('new_password','')
        if not current_user.check_password(cp): flash('Mot de passe actuel incorrect.','error')
        elif len(np) < 6: flash('6 caracteres minimum.','error')
        elif np != request.form.get('confirm_password',''): flash('Mots de passe differents.','error')
        else: current_user.set_password(np); db.session.commit(); log_action('change_password', f'{current_user.full_name} a change son mot de passe'); flash('Mot de passe change !','success'); return redirect(url_for('dashboard'))
    return render_template('change_password.html')

# ═══════ A D M I N   U S E R S ═══════
@app.route('/admin/users')
@permission_required('manage_users')
def manage_users():
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('manage_users.html', users=users, all_roles=CustomRole.query.order_by(CustomRole.name).all())

@app.route('/admin/users/create', methods=['POST'])
@permission_required('manage_users')
def create_user():
    email = request.form.get('email','').strip().lower(); name = request.form.get('full_name','').strip()
    pw = request.form.get('password',''); rid = request.form.get('role_id')
    try: rid = int(rid) if rid else None
    except: rid = None
    if not email or not name or not pw: flash('Tous les champs requis.','error'); return render_template('register.html')
    elif len(pw) < 6: flash('6 caracteres minimum.','error'); return render_template('register.html')
    elif User.query.filter_by(email=email).first(): flash('Email deja utilise.','error')
    else:
        u = User(email=email, full_name=name); u.set_password(pw)
        if rid: u.role_id = rid
        db.session.add(u); db.session.commit()
        notify_user(u.id, 'Compte active', f'Votre compte a ete cree par {current_user.full_name}. Bienvenue !', '/dashboard')
        log_action('create_user', f'Compte cree par admin : {name} ({u.role_name})'); flash(f'Compte cree : {name} ({u.role_name})','success')
    return redirect(url_for('manage_users'))

@app.route('/admin/users/<int:uid>/edit', methods=['POST'])
@permission_required('manage_users')
def edit_user(uid):
    u = db.session.get(User, uid)
    if not u: flash('Introuvable.','error'); return redirect(url_for('manage_users'))
    # Allow editing yourself (just can't change your own role)
    is_self = (u.id == current_user.id)
    u.full_name = request.form.get('full_name',u.full_name).strip()
    new_email = request.form.get('email','').strip().lower()
    if new_email and new_email != u.email and User.query.filter_by(email=new_email).first():
        flash('Cet email est deja utilise.','error'); return redirect(url_for('manage_users'))
    if new_email: u.email = new_email
    if not is_self:
        rid = request.form.get('role_id')
        try: u.role_id = int(rid) if rid else None
        except: pass
    np = request.form.get('new_password','')
    if np and len(np) >= 6: u.set_password(np); flash(f'{u.full_name} mis a jour + mdp change.','success')
    elif np: flash('Mdp non change (6 car. min).','error')
    else: log_action('edit_user', f'Modification du compte de {u.full_name}'); flash(f'{u.full_name} mis a jour.','success')
    db.session.commit()
    return redirect(url_for('manage_users'))

@app.route('/admin/users/<int:uid>/delete', methods=['POST'])
@permission_required('manage_users')
def delete_user(uid):
    u = db.session.get(User, uid)
    if not u: flash('Introuvable.','error')
    elif u.id == current_user.id: flash('Impossible de se supprimer.','error')
    elif Borrow.query.filter_by(user_id=uid).filter(Borrow.status.in_(['active','late'])).first(): flash(f'{u.full_name} a des emprunts en cours.','error')
    else:
        # Nettoyer toutes les tables qui referencent cet utilisateur
        ActivityLog.query.filter_by(user_id=uid).delete()
        Notification.query.filter_by(user_id=uid).delete()
        EventAssignment.query.filter_by(user_id=uid).delete()
        Borrow.query.filter_by(user_id=uid).update({Borrow.user_id: None})
        Event.query.filter_by(created_by=uid).update({Event.created_by: None})
        db.session.delete(u); db.session.commit()
        log_action('delete_user', f'Suppression du compte de {u.full_name}'); flash(f'{u.full_name} supprime.','info')
    return redirect(url_for('manage_users'))

# ═══════ R O L E S ═══════
@app.route('/admin/roles')
@permission_required('manage_roles')
def manage_roles():
    return render_template('manage_roles.html', roles=CustomRole.query.order_by(CustomRole.name).all(), all_permissions=ALL_PERMISSIONS)

@app.route('/admin/roles/create', methods=['POST'])
@permission_required('manage_roles')
def create_role():
    name = request.form.get('name','').strip()
    if not name: flash('Nom requis.','error')
    elif CustomRole.query.filter_by(name=name).first(): flash('Ce role existe deja.','error')
    else:
        r = CustomRole(name=name, icon=request.form.get('icon','👤'), description=request.form.get('description','').strip())
        r.set_permissions(request.form.getlist('permissions')); db.session.add(r); db.session.commit()
        log_action('create_role', f'Role "{name}" cree'); flash(f'Role "{name}" cree.','success')
    return redirect(url_for('manage_roles'))

@app.route('/admin/roles/<int:rid>/edit', methods=['POST'])
@permission_required('manage_roles')
def edit_role(rid):
    r = db.session.get(CustomRole, rid)
    if not r: flash('Role introuvable.','error'); return redirect(url_for('manage_roles'))
    r.name = request.form.get('name',r.name).strip(); r.icon = request.form.get('icon',r.icon)
    r.description = request.form.get('description','').strip()
    r.set_permissions(request.form.getlist('permissions')); db.session.commit()
    log_action('edit_role', f'Role "{r.name}" modifie'); flash(f'Role "{r.name}" mis a jour.','success')
    return redirect(url_for('manage_roles'))

@app.route('/admin/roles/<int:rid>/delete', methods=['POST'])
@permission_required('manage_roles')
def delete_role(rid):
    r = db.session.get(CustomRole, rid)
    if not r: flash('Role introuvable.','error')
    elif User.query.filter_by(role_id=rid).first(): flash('Des utilisateurs utilisent ce role.','error')
    else: db.session.delete(r); db.session.commit(); log_action('delete_role', f'Role "{r.name}" supprime'); flash(f'Role "{r.name}" supprime.','info')
    return redirect(url_for('manage_roles'))

@app.route('/admin/logs/clear', methods=['POST'])
@permission_required('clear_history')
def clear_logs():
    count = ActivityLog.query.count()
    ActivityLog.query.delete()
    db.session.commit()
    db.session.add(ActivityLog(user_id=current_user.id, action='clear_history', description=f'Logs effaces ({count} entrees)'))
    db.session.commit()
    flash(f'{count} logs d\'activite effaces.','success')
    return redirect(url_for('activity_logs'))


@app.route('/admin/import-csv', methods=['GET','POST'])
@permission_required('manage_database')
def import_csv():
    if request.method == 'POST':
        pw = request.form.get('password','')
        if not current_user.check_password(pw):
            flash('Mot de passe incorrect.','error')
            return redirect(url_for('import_csv'))
        table = request.form.get('table','')
        csv_file = request.files.get('csvfile')
        if not csv_file or not table:
            flash('Selectionnez une table et un fichier CSV.','error')
            return redirect(url_for('import_csv'))
        
        # Read CSV (accepte la virgule OU le point-virgule comme separateur)
        import io, csv as csv_module
        raw = csv_file.read().decode('utf-8-sig')
        header = raw.split('\n', 1)[0]
        delim = ';' if header.count(';') > header.count(',') else ','
        reader = csv_module.DictReader(io.StringIO(raw), delimiter=delim)
        rows = list(reader)
        if not rows:
            flash('Fichier CSV vide.','error')
            return redirect(url_for('import_csv'))
        
        model_map = {
            'roles': CustomRole,
            'users': User,
            'categories': Category,
            'equipment': Equipment,
            'equipment_images': EquipmentImage,
            'borrows': Borrow,
            'events': Event,
            'event_assignments': EventAssignment,
            'activity_logs': ActivityLog,
            'notifications': Notification,
        }
        
        model = model_map.get(table)
        if not model:
            flash('Table inconnue.','error')
            return redirect(url_for('import_csv'))
        
        # Detect columns and import
        cols = list(rows[0].keys())
        count = 0
        erreurs = []
        comptes_sans_mdp = []
        for i, row in enumerate(rows, 1):
            try:
                with db.session.begin_nested():
                    # Check if record already exists
                    existing = None
                    if 'id' in cols and row.get('id','').strip():
                        try:
                            existing = db.session.get(model, int(row['id'].strip()))
                        except Exception:
                            existing = None
                    # Fallback: check by unique field (name/email/reference)
                    if not existing:
                        if table == 'roles' and row.get('name','').strip():
                            existing = model.query.filter_by(name=row.get('name','').strip()).first()
                        elif table == 'users' and row.get('email','').strip():
                            existing = model.query.filter_by(email=row.get('email','').strip().lower()).first()
                        elif table == 'categories' and row.get('name','').strip():
                            existing = model.query.filter_by(name=row.get('name','').strip()).first()
                        elif table == 'equipment' and row.get('reference','').strip():
                            existing = model.query.filter_by(reference=row.get('reference','').strip()).first()

                    obj = existing if existing else model()
                    plain_pw = None
                    for col in cols:
                        val = row.get(col, '').strip()
                        if val == '' or val.lower() == 'none':
                            val = None
                            if col == 'id' and existing: continue  # dont overwrite id
                            setattr(obj, col, val)
                            continue
                        # Convert types
                        if col == 'id' and existing: continue
                        if table == 'users' and col == 'password':
                            plain_pw = val
                            continue
                        if col.endswith('_id') or col == 'id' or col.endswith('_quantity'):
                            try: val = int(val)
                            except: pass
                        # Handle date/datetime columns
                        if col in ('created_at', 'uploaded_at', 'borrow_date', 'actual_return_date', 'timestamp'):
                            parsed = None
                            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%d'):
                                try: parsed = datetime.strptime(val, fmt); break
                                except: pass
                            val = parsed
                        if col in ('expected_return_date', 'pickup_date') and val:
                            try: val = datetime.strptime(val, '%Y-%m-%d').date()
                            except:
                                for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
                                    try: val = datetime.strptime(val, fmt).date(); break
                                    except: pass
                                else: val = None
                        if col == 'permissions' and val and not val.startswith('['):
                            val = f'["{val}"]'  # ensure JSON format
                        setattr(obj, col, val)
                    # Utilisateurs: gerer le mot de passe et les champs obligatoires
                    if table == 'users':
                        if obj.email:
                            obj.email = obj.email.strip().lower()
                        if not obj.email:
                            raise ValueError('email vide')
                        if plain_pw:
                            obj.set_password(plain_pw)
                        elif existing is None and not obj.password_hash:
                            obj.set_password('Imagine123')
                            comptes_sans_mdp.append(obj.email)
                        if not obj.full_name:
                            obj.full_name = obj.email.split('@')[0]
                        if obj.role_id and not db.session.get(CustomRole, obj.role_id):
                            obj.role_id = None  # role inconnu -> aucun role
                    if not existing:
                        db.session.add(obj)
                    db.session.flush()
                    count += 1
            except Exception as e:
                erreurs.append(f'Ligne {i}: {str(e)}')
        
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            msg = str(e)
            if 'not-null' in msg.lower() or 'not null' in msg.lower():
                flash('Erreur: une colonne obligatoire est vide (email, nom, mot de passe...). Verifie ton fichier CSV.','error')
            elif 'duplicate' in msg.lower() or 'unique' in msg.lower():
                flash('Erreur: une valeur existe deja (email, nom ou reference duplique).','error')
            elif 'foreign key' in msg.lower():
                flash('Erreur: une reference pointe vers un enregistrement inexistant. Importe d abord la table liee (ex: roles avant users).','error')
            else:
                flash(f'Erreur: {msg}','error')
            return redirect(url_for('import_csv'))
        
        # SECURITE: ne jamais perdre manage_database sur son propre role
        if table == 'roles':
            my_role = db.session.get(CustomRole, current_user.role_id) if current_user.role_id else None
            if my_role and 'manage_database' not in my_role.get_permissions():
                perms = my_role.get_permissions()
                perms.append('manage_database')
                my_role.set_permissions(perms)
                db.session.commit()
                flash(f'{count} ligne(s) importee(s). (Permission manage_database conservee sur votre role.)','success')
            else:
                flash(f'{count} ligne(s) importee(s)/mise(s) a jour dans {table}.','success')
        else:
            flash(f'{count} ligne(s) importee(s)/mise(s) a jour dans {table}.','success')
        if comptes_sans_mdp:
            flash(f'Comptes crees sans mot de passe -> mot de passe temporaire "Imagine123" pour: {", ".join(comptes_sans_mdp)}','warning')
        if erreurs:
            flash(f'{len(erreurs)} ligne(s) ignoree(s): ' + ' | '.join(erreurs[:5]), 'warning')
        return redirect(url_for('import_csv'))
    
    tables = ['roles','users','categories','equipment','equipment_images','borrows','events','event_assignments','activity_logs','notifications']
    return render_template('import_csv.html', tables=tables)


@app.route('/admin/reset-tables', methods=['GET','POST'])
@permission_required('manage_database')
def reset_all_tables():
    if request.method == 'POST':
        pw = request.form.get('password','')
        if not current_user.check_password(pw):
            flash('Mot de passe incorrect.','error')
            return redirect(url_for('reset_all_tables'))
        # Keep current admin
        aid = current_user.id
        rid = current_user.role_id
        Notification.query.delete()
        EventAssignment.query.delete()
        Event.query.delete()
        ActivityLog.query.delete()
        Borrow.query.delete()
        EquipmentImage.query.delete()
        Equipment.query.delete()
        Category.query.delete()
        User.query.filter(User.id != aid).delete()
        CustomRole.query.filter(CustomRole.id != rid).delete()
        db.session.commit()
        flash('Toutes les tables videes (votre compte preserve). Base prete.','success')
        return redirect(url_for('reset_all_tables'))
    return render_template('reset_tables.html')


@app.route('/admin/clear-history', methods=['POST'])
@permission_required('clear_history')
def clear_history():
    if not current_user.check_password(request.form.get('password','')): flash('Mot de passe incorrect.','error'); return redirect(url_for('dashboard'))
    c = Borrow.query.filter_by(status='returned').count()
    Borrow.query.filter_by(status='returned').delete(); db.session.commit()
    log_action('clear_history', f'Historique global efface ({c} emprunts)')
    flash(f'{c} emprunt(s) effaces.','success')
    return redirect(url_for('dashboard'))

@app.route('/equipment/<int:eid>/qrcode')
@login_required
def equipment_qrcode(eid):
    import qrcode, io as bio
    eq = db.session.get(Equipment, eid)
    if not eq:
        return "Introuvable", 404
    url = request.host_url.rstrip('/') + url_for('equipment_detail', eid=eid)
    qr = qrcode.QRCode(box_size=6, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='#0B1D3A', back_color='white')
    buf = bio.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return send_file(buf, mimetype='image/png')


@app.route('/equipment/<int:eid>/qrcode-print')
@login_required
def equipment_qrcode_print(eid):
    eq = db.session.get(Equipment, eid)
    if not eq:
        flash('Introuvable.','error'); return redirect(url_for('dashboard'))
    try:
        qty = int(request.args.get('qty', 1))
    except (TypeError, ValueError):
        qty = 1
    qty = max(1, min(qty, 500))  # entre 1 et 500 etiquettes
    return render_template('equipment_qrcode.html', eq=eq, qty=qty)


@app.route('/qrcodes/all')
@login_required
def all_qrcodes():
    page = request.args.get('page', 1, type=int)
    per_page = 15
    equipment_list = Equipment.query.order_by(Equipment.name).all()
    total = len(equipment_list)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(max(1, page), total_pages)
    start = (page - 1) * per_page
    page_equipment = equipment_list[start:start+per_page]
    return render_template('all_qrcodes.html', equipment=page_equipment, page=page, total_pages=total_pages, total=total)


@app.route('/equipment/<int:eid>/clear-history', methods=['POST'])
@permission_required('clear_history')
def clear_equipment_history(eid):
    eq = db.session.get(Equipment, eid)
    if not eq: flash('Introuvable.','error'); return redirect(url_for('dashboard'))
    if not current_user.check_password(request.form.get('password','')): flash('Mot de passe incorrect.','error'); return redirect(url_for('equipment_detail', eid=eid))
    c = Borrow.query.filter_by(equipment_id=eid, status='returned').count()
    Borrow.query.filter_by(equipment_id=eid, status='returned').delete(); db.session.commit()
    log_action('clear_history', f'Historique efface pour {eq.name} ({c} emprunts)'); flash(f'{c} emprunt(s) effaces pour {eq.name}.','success')
    return redirect(url_for('equipment_detail', eid=eid))

@app.route('/admin/logs')
@permission_required('view_logs')
def activity_logs():
    page = request.args.get('page',1,type=int)
    action_filter = request.args.get('action','').strip()
    user_filter = request.args.get('user','').strip()
    q = ActivityLog.query
    if action_filter: q = q.filter_by(action=action_filter)
    if user_filter:
        # Search by user name
        matching_users = User.query.filter(User.full_name.ilike(f'%{user_filter}%')).all()
        uids = [u.id for u in matching_users]
        if uids: q = q.filter(ActivityLog.user_id.in_(uids))
        else: q = q.filter(ActivityLog.user_id == -1)  # no results
    logs = q.order_by(ActivityLog.timestamp.desc()).paginate(page=page, per_page=50, error_out=False)
    # Stats
    total_today = ActivityLog.query.filter(db.func.date(ActivityLog.timestamp) == db.func.date(tunisia_now())).count()
    total_week = ActivityLog.query.filter(ActivityLog.timestamp >= tunisia_now() - timedelta(days=7)).count()
    all_actions = [r[0] for r in db.session.query(ActivityLog.action).distinct().order_by(ActivityLog.action).all()]
    ACTION_LABELS = {
        'login':'🔑 Connexion','logout':'🚪 Deconnexion','failed_login':'⚠️ Echec connexion',
        'register':'📝 Inscription','borrow':'📤 Emprunt','return':'✅ Retour',
        'add_equipment':'➕ Ajout materiel','edit_equipment':'✏️ Modif materiel',
        'delete_equipment':'🗑️ Suppr materiel','add_category':'📂 Categorie ajoutee',
        'delete_category':'❌ Categorie supprimee','create_user':'👤 Compte cree',
        'edit_user':'✏️ Compte modifie','delete_user':'❌ Compte supprime',
        'create_role':'🔐 Role cree','edit_role':'✏️ Role modifie','delete_role':'❌ Role supprime',
        'change_password':'🔑 Mdp change','create_event':'📅 Evenement cree',
        'edit_event':'✏️ Evenement modifie','delete_event':'🗑️ Evenement supprime',
        'clear_past_events':'🧹 Evenements passes effaces','clear_history':'🧹 Effacement historique',
        'export_excel':'📊 Export Excel','permission_denied':'🚫 Acces refuse',
    }
    return render_template('activity_logs.html', logs=logs, action_filter=action_filter, user_filter=user_filter, total_today=total_today, total_week=total_week, all_actions=all_actions, ACTION_LABELS=ACTION_LABELS, all_users=User.query.order_by(User.full_name).all())

@app.route('/categories/add', methods=['POST'])
@permission_required('manage_categories')
def add_category():
    name = request.form.get('name','').strip(); icon = request.form.get('icon','📦')
    if name and not Category.query.filter_by(name=name).first():
        db.session.add(Category(name=name, icon=icon)); db.session.commit()
        log_action('add_category', f'Categorie "{name}" ajoutee'); flash(f'Categorie "{name}" ajoutee.','success')
    return redirect(url_for('dashboard'))

@app.route('/categories/<int:cid>/delete', methods=['POST'])
@permission_required('manage_categories')
def delete_category(cid):
    cat = db.session.get(Category, cid)
    if not cat: flash('Categorie introuvable.','error')
    elif Equipment.query.filter_by(category_id=cid).first(): flash(f'Des materiels utilisent la categorie {cat.name}.','error')
    else: db.session.delete(cat); db.session.commit(); log_action('delete_category', f'Categorie "{cat.name}" supprimee'); flash(f'Categorie "{cat.name}" supprimee.','info')
    return redirect(url_for('dashboard'))

# ═══════════ N E W :  E M P L O I   D U   T E M P S ═══════════
@app.route('/schedule')
@login_required
def schedule():
    if is_pending_user(): flash('Votre compte est en attente de validation.','error'); return redirect(url_for('dashboard'))
    """Main calendar/schedule view"""
    past = Event.query.filter(Event.end_date < date.today()).order_by(Event.event_date.desc()).limit(20).all()
    # Tous les evenements (passes + a venir) pour le calendrier mensuel
    all_events = Event.query.order_by(Event.event_date, Event.start_time).all()
    events_json = json.dumps([
        {'id': e.id, 'title': e.title, 'description': e.description or '',
         'date': e.event_date.strftime('%Y-%m-%d'), 'dend': e.date_end().strftime('%Y-%m-%d'),
         'start': e.start_time or '08:00', 'end': e.end_time or '17:00',
         'location': e.location or '', 'status': e.status,
         'users': [a.user_id for a in e.assignments]}
        for e in all_events
    ])
    # URL publique du flux ICS (Google Calendar / Apple / Outlook)
    site = request.host_url.rstrip('/')
    if not site.startswith('http://localhost') and not site.startswith('http://127.0.0.1'):
        site = site.replace('http://', 'https://', 1)
    ics_url = site + '/calendar.ics?key=' + ICS_KEY
    all_users = User.query.order_by(User.full_name).all()
    user_names = dict([(u.id, u.full_name) for u in all_users])  # pour le JS (pas de comprehension Jinja)
    return render_template('schedule.html', past=past, today=tunisia_now().date(), all_users=all_users, events_json=events_json, ics_url=ics_url, user_names=user_names, today_str=tunisia_now().date().strftime('%Y-%m-%d'))

# ── Flux ICS : synchronisation avec Google Calendar, Apple Calendar, Outlook... ──
def ics_escape(text):
    if not text: return ''
    return text.replace('\\', '\\\\').replace(';', '\\;').replace(',', '\\,').replace('\n', '\\n')

def ics_fold(lines):
    """Plie les lignes > 75 octets (regle RFC 5545) sans couper un caractere unicode."""
    out = []
    for ln in lines:
        b = ln.encode('utf-8')
        if len(b) <= 75:
            out.append(ln); continue
        parts = []
        while b:
            chunk = b[:75]
            while chunk and (chunk[-1] & 0xC0) == 0x80:
                chunk = chunk[:-1]
            parts.append(chunk)
            b = b[len(chunk):]
        out.append(parts[0].decode('utf-8'))
        for p in parts[1:]:
            out.append(' ' + p.decode('utf-8'))
    return out

@app.route('/calendar.ics')
def calendar_ics():
    """Flux iCal public (protege par une cle derivee de SECRET_KEY).
    A ajouter dans Google Calendar (Autres calendriers -> Ajouter un calendrier
    -> A partir d'une URL) : Google le recupere automatiquement et met a jour
    les evenements (creations, modifications, annulations).
    Param ?dl=1 : telechargement du fichier .ics a la main."""
    if request.args.get('key') != ICS_KEY:
        return 'Cle de souscription invalide.', 403
    today = date.today()
    horizon = today + timedelta(days=92)
    events = Event.query.filter(Event.end_date >= today, Event.event_date <= horizon).order_by(Event.event_date, Event.start_time).all()
    dtstamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
    lines = [
        'BEGIN:VCALENDAR', 'VERSION:2.0',
        'PRODID:-//Imagine Events Tunisia//Planning Depot//FR',
        'CALSCALE:GREGORIAN', 'METHOD:PUBLISH', 'PUBLISHED:' + dtstamp,
        'X-WR-CALNAME:Imagine Events - Planning', 'X-WR-TIMEZONE:Africa/Tunis',
        'BEGIN:VTIMEZONE', 'TZID:Africa/Tunis', 'BEGIN:STANDARD',
        'DTSTART:19700101T000000', 'TZOFFSETFROM:+0100', 'TZOFFSETTO:+0100',
        'TZNAME:CET', 'END:STANDARD', 'END:VTIMEZONE',
    ]
    for evt in events:
        d = evt.event_date.strftime('%Y%m%d')
        d_end = evt.date_end().strftime('%Y%m%d')  # multi-jours : derniere journee
        st = (evt.start_time or '08:00').replace(':', '') + '00'
        en = (evt.end_time or '17:00').replace(':', '') + '00'
        assignees = ', '.join(a.user.full_name for a in evt.assignments if a.user)
        desc = (evt.description or '').strip()
        if assignees:
            desc = (desc + '\n' if desc else '') + 'Equipe: ' + assignees
        lines += [
            'BEGIN:VEVENT',
            'UID:imagine-events-' + str(evt.id) + '@imagine-events.tn',
            'DTSTAMP:' + dtstamp,
            'DTSTART;TZID=Africa/Tunis:' + d + 'T' + st,
            'DTEND;TZID=Africa/Tunis:' + d_end + 'T' + en,
            'SUMMARY:' + ics_escape(evt.title),
            'DESCRIPTION:' + ics_escape(desc),
        ]
        if evt.location:
            lines.append('LOCATION:' + ics_escape(evt.location))
        lines.append('STATUS:' + ('CANCELLED' if evt.status == 'cancelled' else 'CONFIRMED'))
        lines.append('END:VEVENT')
    lines.append('END:VCALENDAR')
    body = '\r\n'.join(ics_fold(lines)) + '\r\n'
    disp = 'attachment' if request.args.get('dl') else 'inline'
    return Response(body, mimetype='text/calendar; charset=utf-8',
                    headers={'Content-Disposition': disp + '; filename="imagine-events-planning.ics"',
                             'Cache-Control': 'public, max-age=3600'})

@app.route('/schedule/create', methods=['POST'])
@permission_required('manage_schedule')
def create_event():
    title = request.form.get('title','').strip()
    if not title: flash('Titre requis.','error'); return redirect(url_for('schedule'))
    try: ed = datetime.strptime(request.form.get('event_date',''),'%Y-%m-%d').date()
    except: flash('Date invalide.','error'); return redirect(url_for('schedule'))
    # ── Evenement multi-jours : date de fin (defaut = le jour de debut) ──
    try:
        ed_end = datetime.strptime(request.form.get('end_date','') or ed.strftime('%Y-%m-%d'),'%Y-%m-%d').date()
    except (TypeError, ValueError):
        ed_end = ed
    if ed_end < ed:
        flash('La date de fin doit etre apres (ou egale a) la date de debut.','error'); return redirect(url_for('schedule'))
    st = request.form.get('start_time','08:00'); et = request.form.get('end_time','17:00')
    evt = Event(title=title, description=request.form.get('description','').strip(), event_date=ed, end_date=ed_end, start_time=st, end_time=et, location=request.form.get('location','').strip(), created_by=current_user.id)
    db.session.add(evt); db.session.flush()
    # Assign users
    user_ids = request.form.getlist('assigned_users')
    for uid in user_ids:
        db.session.add(EventAssignment(event_id=evt.id, user_id=int(uid), role='Staff'))
    db.session.commit()
    # Notify assigned users
    for uid in user_ids:
        u = db.session.get(User, uid)
        if u:
            notify_user(u.id, f'Nouvel evenement : {title}', f'Vous etes assigne a "{title}" {evt.date_label()} ({st}-{et})', '/schedule')
    log_action('create_event', f'Evenement "{title}" cree {evt.date_label()}'); flash(f'Evenement "{title}" cree {evt.date_label()}.','success')
    return redirect(url_for('schedule'))

@app.route('/schedule/<int:evid>/edit', methods=['POST'])
@permission_required('manage_schedule')
def edit_event(evid):
    evt = db.session.get(Event, evid)
    if not evt: flash('Introuvable.','error'); return redirect(url_for('schedule'))
    evt.title = request.form.get('title','').strip(); evt.description = request.form.get('description','').strip()
    try: evt.event_date = datetime.strptime(request.form.get('event_date',''),'%Y-%m-%d').date()
    except: pass
    # ── Date de fin (evenement multi-jours) ; vide = 1 jour ──
    end_form = request.form.get('end_date','').strip()
    if end_form:
        try:
            new_end = datetime.strptime(end_form, '%Y-%m-%d').date()
            evt.end_date = new_end if new_end >= evt.event_date else evt.event_date
        except (TypeError, ValueError):
            evt.end_date = evt.event_date
    else:
        evt.end_date = evt.event_date
    evt.start_time = request.form.get('start_time','08:00'); evt.end_time = request.form.get('end_time','17:00')
    evt.location = request.form.get('location','').strip(); evt.status = request.form.get('status',evt.status)
    # Update assignments
    EventAssignment.query.filter_by(event_id=evid).delete()
    for uid in request.form.getlist('assigned_users'):
        db.session.add(EventAssignment(event_id=evid, user_id=int(uid)))
    db.session.commit(); log_action('edit_event', f'Evenement "{evt.title}" modifie'); flash(f'Evenement modifie.','success')
    return redirect(url_for('schedule'))

@app.route('/schedule/<int:evid>/delete', methods=['POST'])
@permission_required('manage_schedule')
def delete_event(evid):
    evt = db.session.get(Event, evid)
    if evt: db.session.delete(evt); db.session.commit(); log_action('delete_event', f'Evenement "{evt.title}" supprime'); flash(f'Evenement "{evt.title}" supprime.','info')
    return redirect(url_for('schedule'))


@app.route('/schedule/clear-past', methods=['POST'])
@permission_required('manage_schedule')
def clear_past_events():
    count = Event.query.filter(Event.end_date < date.today()).count()
    past_ids = [e.id for e in Event.query.filter(Event.end_date < date.today()).all()]
    EventAssignment.query.filter(EventAssignment.event_id.in_(past_ids)).delete(synchronize_session='fetch')
    Event.query.filter(Event.end_date < date.today()).delete()
    db.session.commit()
    log_action('clear_past_events', f'{count} evenements passes supprimes'); flash(f'{count} evenement(s) passes supprime(s).','success')
    return redirect(url_for('schedule'))

@app.route('/schedule/<int:evid>/assign', methods=['POST'])
@permission_required('manage_schedule')
def assign_one_user(evid):
    """Quick assign one user from schedule page"""
    uid = request.form.get('user_id'); role = request.form.get('role','Staff')
    if uid and not EventAssignment.query.filter_by(event_id=evid, user_id=int(uid)).first():
        db.session.add(EventAssignment(event_id=evid, user_id=int(uid), role=role)); db.session.commit()
    return redirect(url_for('schedule'))

@app.route('/schedule/<int:evid>/unassign/<int:uid>', methods=['POST'])
@permission_required('manage_schedule')
def unassign_user(evid, uid):
    ea = EventAssignment.query.filter_by(event_id=evid, user_id=uid).first()
    if ea: db.session.delete(ea); db.session.commit()
    return redirect(url_for('schedule'))

# ═══════════ N E W :  N O T I F I C A T I O N S ═══════════
@app.route('/notifications')
@login_required
def notifications():
    if is_pending_user(): return redirect(url_for('dashboard'))
    notifs = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).limit(100).all()
    return render_template('notifications.html', notifs=notifs)

@app.route('/notifications/read/<int:nid>', methods=['POST'])
@login_required
def mark_read(nid):
    n = Notification.query.filter_by(id=nid, user_id=current_user.id).first()
    if n: n.read = True; db.session.commit()
    return redirect(request.form.get('redirect','/notifications'))

@app.route('/notifications/read-all', methods=['POST'])
@login_required
def mark_all_read():
    Notification.query.filter_by(user_id=current_user.id, read=False).update({Notification.read: True})
    db.session.commit(); flash('Toutes les notifications lues.','success')
    return redirect(url_for('notifications'))


@app.route('/notifications/delete/<int:nid>', methods=['POST'])
@login_required
def delete_notification(nid):
    n = Notification.query.filter_by(id=nid, user_id=current_user.id).first()
    if n: db.session.delete(n); db.session.commit()
    return redirect(url_for('notifications'))


@app.route('/notifications/clear-all', methods=['POST'])
@login_required
def clear_all_notifications():
    count = Notification.query.filter_by(user_id=current_user.id).count()
    Notification.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()
    flash(f'{count} notification(s) supprimees.','success')
    return redirect(url_for('notifications'))


# ═══════════ N E W :  E X P O R T S ═══════════
@app.route('/export/photos-zip')
@permission_required('manage_database')
def export_photos_zip():
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 1) Photos en base de donnees (source de verite)
        for b in ImageBlob.query.all():
            zf.writestr(b.filename, b.data or b'')
        # 2) Fichiers sur le disque pas encore en base
        upload_dir = app.config['UPLOAD_FOLDER']
        for fn in os.listdir(upload_dir):
            if fn == '.gitkeep': continue
            fpath = os.path.join(upload_dir, fn)
            if os.path.isfile(fpath) and not ImageBlob.query.filter_by(filename=fn).first():
                zf.write(fpath, fn)
    buf.seek(0)
    return send_file(buf, mimetype='application/zip', as_attachment=True, download_name=f'photos_imagine_{date.today().strftime("%Y%m%d")}.zip')


@app.route('/admin/photos/save-to-db', methods=['POST'])
@permission_required('manage_database')
def save_photos_to_db():
    """Copie toutes les photos actuellement sur le disque dans la base de donnees."""
    added = 0; skipped = 0
    for fn in os.listdir(app.config['UPLOAD_FOLDER']):
        fp = os.path.join(app.config['UPLOAD_FOLDER'], fn)
        if not os.path.isfile(fp) or fn == '.gitkeep': continue
        if ImageBlob.query.filter_by(filename=fn).first():
            skipped += 1; continue
        try:
            with open(fp, 'rb') as fh:
                data = fh.read()
            low = fn.lower()
            mt = 'image/png' if low.endswith('.png') else 'image/jpeg' if low.endswith(('.jpg','.jpeg')) else 'image/gif' if low.endswith('.gif') else 'application/octet-stream'
            db.session.add(ImageBlob(filename=fn, data=data, mimetype=mt))
            added += 1
        except Exception as e:
            print(f'[IMG-DB] echec {fn}: {e}')
    db.session.commit()
    flash(f'{added} photo(s) sauvegardee(s) dans la base. {skipped} deja presente(s). Elles ne disparaitront plus aux mises a jour.','success')
    return redirect(url_for('restore_photos'))


@app.route('/admin/restore-photos', methods=['GET','POST'])
@permission_required('manage_database')
def restore_photos():
    if request.method == 'POST':
        zf = request.files.get('zipfile')
        if not zf or not zf.filename.endswith('.zip'):
            flash('Veuillez selectionner un fichier ZIP valide.','error')
            return redirect(url_for('restore_photos'))
        import zipfile
        count = 0; en_base = 0
        with zipfile.ZipFile(zf) as z:
            for fn in z.namelist():
                if fn.endswith('/') or fn == '.gitkeep': continue
                base = os.path.basename(fn)
                data = z.read(fn)
                if not data: continue
                # 1) Stocker dans la base de donnees (permanent)
                if not ImageBlob.query.filter_by(filename=base).first():
                    low = base.lower()
                    mt = 'image/png' if low.endswith('.png') else 'image/jpeg' if low.endswith(('.jpg','.jpeg')) else 'image/gif' if low.endswith('.gif') else 'application/octet-stream'
                    db.session.add(ImageBlob(filename=base, data=data, mimetype=mt))
                    en_base += 1
                # 2) Aussi sur le disque (comme avant)
                target = os.path.join(app.config['UPLOAD_FOLDER'], base)
                if not os.path.exists(target):
                    with open(target, 'wb') as fh:
                        fh.write(data)
                count += 1
        db.session.commit()
        flash(f'{count} photo(s) restauree(s), dont {en_base} sauvegardee(s) en base. Elles ne disparaitront plus aux mises a jour.','success')
        return redirect(url_for('restore_photos'))
    return render_template('restore_photos.html')


@app.route('/export/excel')
@permission_required('manage_equipment')
def export_excel():
    log_action('export_excel', 'Export Excel de l\'inventaire')
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Inventaire Imagine Events"
    # Styles
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0B1D3A', end_color='0B1D3A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    green_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    red_fill = PatternFill(start_color='FDE8EC', end_color='FDE8EC', fill_type='solid')
    # Title
    ws.merge_cells('A1:I1'); ws['A1'] = 'IMAGINE EVENTS TUNISIA - Inventaire du Depot'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='0B1D3A'); ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:I2'); ws['A2'] = f'Exporte le {date.today().strftime("%d/%m/%Y")}'
    ws['A2'].font = Font(name='Arial', size=9, color='64748B'); ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([])  # blank row
    # Headers
    headers = ['Reference', 'Nom', 'Categorie', 'Description', 'Qté Totale', 'Qté Dispo', 'État', 'Emplacement', 'Spécifications']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment; cell.border = thin_border
    # Data
    for r, eq in enumerate(Equipment.query.order_by(Equipment.name).all(), 5):
        cat_name = eq.category.name if eq.category else ''
        data = [eq.reference, eq.name, cat_name, eq.description, eq.total_quantity, eq.available_quantity, eq.condition, eq.location, eq.specifications.replace('\n',' | ')[:200]]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = cell_font; cell.border = thin_border
            if c in (5,6): cell.alignment = Alignment(horizontal='center')
        # Color rows
        fill = green_fill if eq.available_quantity > 0 else red_fill
        for c in range(1, 10): ws.cell(row=r, column=c).fill = fill
    # Column widths
    widths = [15, 28, 18, 30, 12, 12, 14, 18, 35]
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'inventaire_imagine_{date.today().strftime("%Y%m%d")}.xlsx')

@app.route('/export/pdf')
@permission_required('manage_equipment')
def export_printer():
    """Page optimisee pour impression (CTRL+P -> PDF)"""
    return render_template('export_print.html', equipment=Equipment.query.order_by(Equipment.name).all(), now=tunisia_now())

# ═══════ I N I T   D B ═══════
def init_db():
    """Cree les tables et les donnees par defaut manquantes.
    Ne plante JAMAIS au demarrage : si un conflit survient (ex: deux demarrages
    simultanes sur Render, ou des donnees deja presentes), on l'ignore et on continue."""
    with app.app_context():
        db.create_all()
        try:
            # ── MIGRATION : ajouter les nouvelles colonnes si elles manquent ──
            def ensure_column(table, col, ddl):
                try:
                    cols = [c['name'] for c in db.inspect(db.engine).get_columns(table)]
                    if col not in cols:
                        with db.engine.begin() as conn:
                            conn.execute(db.text(ddl))
                        print(f'[INIT] colonne {col} ajoutee a la table {table}')
                except Exception as e:
                    print(f'[INIT] migration {table}.{col} ignoree: {str(e)[:120]}')
            ensure_column('equipment', 'in_repair', 'ALTER TABLE equipment ADD COLUMN in_repair INTEGER DEFAULT 0')
            ensure_column('borrows', 'event_id', 'ALTER TABLE borrows ADD COLUMN event_id INTEGER')
            ensure_column('borrows', 'pickup_date', 'ALTER TABLE borrows ADD COLUMN pickup_date DATE')
            ensure_column('events', 'end_date', 'ALTER TABLE events ADD COLUMN end_date DATE')
            # Backfill : les evenements existants (1 jour) -> end_date = event_date
            # (les requetes filtrent ensuite simplement sur end_date)
            try:
                db.session.execute(db.text('UPDATE events SET end_date = event_date WHERE end_date IS NULL'))
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print('[INIT] backfill events.end_date ignore:', str(e)[:120])

            # ── Roles par defaut (cree seulement s'ils manquent) ──
            def ensure_role(name, icon, description, perms):
                r = CustomRole.query.filter_by(name=name).first()
                if not r:
                    r = CustomRole(name=name, icon=icon, description=description)
                    r.set_permissions(perms)
                    db.session.add(r)
                return r

            ar = ensure_role('Admin', '👑', 'Toutes les permissions', [p['key'] for p in ALL_PERMISSIONS])
            ensure_role('Staff', '👷', 'Emprunts et retours', ['borrow_equipment', 'return_equipment'])
            ensure_role('Manager', '🛡️', 'Gestion complete sans effacer historique', ['manage_users','manage_equipment','borrow_equipment','return_equipment','manage_categories','view_logs','manage_schedule'])
            ensure_role('En attente', '⏳', 'Nouveau compte en attente', [])
            db.session.flush()  # pour obtenir l'id du role Admin

            # ── Admin par defaut (toujours present) ──
            admin = User.query.filter_by(email='admin@imagine-events.com').first()
            if not admin:
                admin = User(email='admin@imagine-events.com', full_name='Admin Imagine', role_id=ar.id)
                admin.set_password('admin123')
                db.session.add(admin)
                print('OK - compte admin cree/recree (admin@imagine-events.com / admin123)')
            elif not admin.role_id:
                # Securite: si le role du compte admin a ete perdu (ex: import CSV), le rattacher
                admin.role_id = ar.id

            # ── Staff par defaut ──
            if not User.query.filter_by(email='staff@imagine-events.com').first():
                sr = CustomRole.query.filter_by(name='Staff').first()
                staff = User(email='staff@imagine-events.com', full_name='Equipe Logistique', role_id=sr.id if sr else None)
                staff.set_password('staff123')
                db.session.add(staff)

            # ── Categories par defaut ──
            cat_names = ['Ecrans & Affichage', 'Sonorisation', 'Eclairage', 'Scenes & Structures', 'Mobilier', 'Cablage & Connectique']
            cat_icons = ['📺', '🎤', '💡', '🎭', '🛋️', '🔌']
            for nm, ic in zip(cat_names, cat_icons):
                if not Category.query.filter_by(name=nm).first():
                    db.session.add(Category(name=nm, icon=ic))
            db.session.commit()

            # ── Materiel d'exemple (seulement si la table est vide) ──
            if Equipment.query.count() == 0:
                cat = {nm: Category.query.filter_by(name=nm).first() for nm in cat_names}
                eqs = [
                    Equipment(name='Ecran LED 43"', description='Ecran LED haute definition 43 pouces.', reference='IM-LED43-001', category_id=cat['Ecrans & Affichage'].id if cat['Ecrans & Affichage'] else None, total_quantity=10, available_quantity=10, specifications='Taille: 43"\nResolution: Full HD\nLuminosite: 350 cd/m2\nConnectique: HDMI, VGA', condition='Excellent', location='Allee A - Etagere 1'),
                    Equipment(name='Ecran LED 55"', description='Ecran LED 55 pouces pour evenements.', reference='IM-LED55-002', category_id=cat['Ecrans & Affichage'].id if cat['Ecrans & Affichage'] else None, total_quantity=6, available_quantity=6, specifications='Taille: 55"\nResolution: 4K UHD', condition='Excellent', location='Allee A - Etagere 2'),
                    Equipment(name='Micro Sans Fil SM58', description='Micro professionnel Shure.', reference='IM-MIC-003', category_id=cat['Sonorisation'].id if cat['Sonorisation'] else None, total_quantity=12, available_quantity=12, specifications='Marque: Shure\nType: Dynamique\nPortee: 100m', condition='Bon etat', location='Allee B - Armoire 1'),
                    Equipment(name='Lyre LED Beam', description='Projecteur lyre motorise a LED.', reference='IM-LYR-004', category_id=cat['Eclairage'].id if cat['Eclairage'] else None, total_quantity=8, available_quantity=8, specifications='LED: 200W\nDMX: 16 canaux\nPrisme: 8 facettes', condition='Excellent', location='Allee C - Etagere 3'),
                    Equipment(name='Canape Lounge Design', description='Canape 3 places pour receptions.', reference='IM-CAN-005', category_id=cat['Mobilier'].id if cat['Mobilier'] else None, total_quantity=15, available_quantity=15, specifications='Places: 3\nMateriau: Velours premium', condition='Tres bon etat', location='Zone Mobilier - Rangee D'),
                    Equipment(name='Barre LED RGB', description='Barre LED 144 LEDs/m pour eclairage.', reference='IM-BAR-006', category_id=cat['Eclairage'].id if cat['Eclairage'] else None, total_quantity=20, available_quantity=20, specifications='LEDs: RGB 144/m\nAngle: 120 degres', condition='Bon etat', location='Allee C - Etagere 1'),
                    Equipment(name='Enceinte Active 15"', description='Enceinte 15" 1000W pour sonorisation.', reference='IM-ENC-007', category_id=cat['Sonorisation'].id if cat['Sonorisation'] else None, total_quantity=8, available_quantity=8, specifications='Puissance: 1000W\nSPL Max: 132 dB', condition='Excellent', location='Allee B - Sol'),
                    Equipment(name='Scene Modulable 2x1m', description='Element de scene aluminium.', reference='IM-SCN-008', category_id=cat['Scenes & Structures'].id if cat['Scenes & Structures'] else None, total_quantity=30, available_quantity=30, specifications='Dimensions: 200x100 cm\nCharge max: 750 kg/m2', condition='Bon etat', location='Zone Scenes'),
                ]
                db.session.add_all(eqs)
                db.session.commit()

            # ── Migration photos : copier en base les photos presentes sur le disque ──
            try:
                for fn in os.listdir(app.config['UPLOAD_FOLDER']):
                    fp = os.path.join(app.config['UPLOAD_FOLDER'], fn)
                    if not os.path.isfile(fp) or fn == '.gitkeep': continue
                    if ImageBlob.query.filter_by(filename=fn).first(): continue
                    with open(fp, 'rb') as fh:
                        data = fh.read()
                    low = fn.lower()
                    mt = 'image/png' if low.endswith('.png') else 'image/jpeg' if low.endswith(('.jpg','.jpeg')) else 'image/gif' if low.endswith('.gif') else 'application/octet-stream'
                    db.session.add(ImageBlob(filename=fn, data=data, mimetype=mt))
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print('[INIT] migration photos ignoree:', type(e).__name__, str(e)[:120])

        except IntegrityError as e:
            db.session.rollback()
            print('[INIT] Conflit au demarrage (base deja initialisee) - ignore:', str(e)[:160])
        except Exception as e:
            db.session.rollback()
            print('[INIT] Avertissement - initialisation ignoree:', type(e).__name__, str(e)[:160])
        print('OK - base prete (admin@imagine-events.com / admin123)')

init_db()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
